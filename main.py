import logging
import multiprocessing
import os
import unittest
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from tests.authentication_test.test_login import TestLogin
from tests.authentication_test.test_register import TestRegister
#from tests.authentication_test.test_setting import TestSetting
#from tests.authentication_test.test_forgotpassword import TestPassword
from tests.deposit_test.test_quickreload import TestQuickReload
from tests.deposit_test.test_banktransfer import TestBankTransfer
from tests.withdraw_test.test_withdrawtransfer import TestWithdrawTransfer
#from tests.withdraw_test.test_ewallet import TestEWallet
#from tests.coupon_test.test_coupon import TestCoupon
#from tests.transaction_history_test.test_history import TestHistory
#from tests.test_profile import TestProfile
from tests.deposit_test.test_spamdeposit import TestSpamDeposit
from tests.transfer_test.test_transfer import TestTransfer
from tests.transfer_test.test_main_provider import TestMainProvider
from tests.transfer_test.test_provider_to_provider import TestProviderToProvider
from tests.revert_test.revert_test import TestRevert


class CustomTestResult(unittest.TextTestResult):

    def __init__(self, stream, descriptions, verbosity):
        super().__init__(stream, descriptions, verbosity)
        self.successes = []
        self.errors_and_failures = []
        self.test_results = []

    def addSuccess(self, test):
        super().addSuccess(test)
        self.successes.append(test)
        self.test_results.append(("PASS", test, None))

    def addError(self, test, err):
        super().addError(test, err)
        self.errors_and_failures.append(("ERROR", test, err))
        self.test_results.append(("ERROR", test, err))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        self.errors_and_failures.append(("FAILURE", test, err))
        self.test_results.append(("FAILURE", test, err))


class CustomTestRunner(unittest.TextTestRunner):

    def __init__(self, language, browser, **kwargs):
        super().__init__(**kwargs)
        self.language = language
        self.browser = browser

    def _makeResult(self):
        return CustomTestResult(self.stream, self.descriptions, self.verbosity)

    def run(self, test):
        result = super().run(test)
        logging.info(f"Total tests run: {result.testsRun}")
        logging.info(f"Successes: {len(result.successes)}")
        logging.info(f"Failures: {len(result.failures)}")
        logging.info(f"Errors: {len(result.errors)}")
        self.write_to_excel(result)
        return result

    def write_to_excel(self, result):
        test_results_by_class = {}
        for status, test, error in result.test_results:
            test_class = test.__class__.__name__
            if test_class not in test_results_by_class:
                test_results_by_class[test_class] = []
            test_results_by_class[test_class].append((status, test, error))

        for test_class, class_results in test_results_by_class.items():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{test_class}.xlsx"

            current_dir = os.path.dirname(os.path.abspath(__file__))
            results_dir = os.path.join(current_dir, "test_results")
            os.makedirs(results_dir, exist_ok=True)

            filepath = os.path.join(results_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = f"{test_class}_{self.language}_{self.browser}"

            pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            header_font = Font(bold=True)

            headers = ["Test Class", "Test Name", "Status", "Error Message"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font

            row = 2
            for status, test, error in class_results:
                ws.cell(row=row, column=1, value=test_class)
                ws.cell(row=row, column=2, value=test._testMethodName)
                ws.cell(row=row, column=3, value=status)

                if error:
                    error_type, error_value, traceback = error
                    error_msg = str(error_value)

                    if isinstance(error_value, AssertionError) or "Test failed:" in error_msg:
                        ws.cell(row=row, column=4, value=error_msg)
                    else:
                        formatted_msg = f"Test failed: {error_msg}"
                        ws.cell(row=row, column=4, value=formatted_msg)

                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = fail_fill
                else:
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = pass_fill

                row += 1

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=4)
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    ws.row_dimensions[row_idx].height = 15 * max(lines, 1)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filepath)
            logging.info(f"Test results Excel file created at: {filepath}")

            if os.path.exists(filepath):
                logging.info(f"Excel file successfully created: {filename}")
            else:
                logging.error(f"Failed to create Excel file: {filename}")


def create_test_suite(language, browser):
    suite = unittest.TestSuite()

    # test_classes = [
    #    TestLogin, TestRegister, TestSetting, TestPassword, TestProfile, TestQuickReload, TestBankTransfer, TestTransfer,
    #    TestSpamDeposit, TestWithdrawTransfer, TestEWallet, TestCoupon, TestHistory,
    #    TestDailyCheckIn, TestDailyMission, TestInviteFriends, TestLuckyWheelSpinPage, TestProfilePage, TestPromotion
    # ]
    test_classes = [TestRevert]

    for test_class in test_classes:
        # Generate test methods if it's TestRevert
        if test_class == TestRevert:
            # Generate the test methods
            TestRevert.generate_test_methods(language=language, browser=browser)
            # Get all test methods (including auto-generated ones)
            test_methods = TestRevert.get_test_methods(language=language, browser=browser)

            # Add each test method to suite
            for method_name in test_methods:
                suite.addTest(test_class(method_name, language=language, browser=browser))
        else:
            # Handle other test classes normally
            tests = unittest.TestLoader().loadTestsFromTestCase(test_class)
            for test in tests:
                suite.addTest(test_class(test._testMethodName, language=language, browser=browser))

    return suite


def run_tests(language, browser):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"TestResults_{language}_{browser}_{timestamp}.log"

    logging.basicConfig(
        level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
            logging.FileHandler(log_filename),
            #logging.StreamHandler()  # This will print to console
        ]
    )

    logging.info(f"Starting test run in {browser} browser for {language} language...")
    suite = create_test_suite(language, browser)
    runner = CustomTestRunner(language, browser, verbosity=2)
    result = runner.run(suite)

    logging.info("\n" + "=" * 50)
    logging.info("TEST RUN SUMMARY")
    logging.info("=" * 50)
    logging.info(f"Total tests run: {result.testsRun}")
    logging.info(f"Successes: {len(result.successes)}")
    logging.info(f"Failures: {len(result.failures)}")
    logging.info(f"Errors: {len(result.errors)}")
    logging.info("=" * 50 + "\n")

    return result.wasSuccessful()


if __name__ == "__main__":
    # languages = ["cn", "en", "bm"]
    #browsers = ["chrome", "firefox","edge","safari"]
    languages = ["bm"]
    browsers = ["firefox"]
    processes = []

    for browser in browsers:
        for language in languages:
            logging.info(f"Starting test run for language: {language}, browser: {browser}")
            process = multiprocessing.Process(target=run_tests, args=(language, browser))
            process.start()
            processes.append(process)

    for process in processes:
        process.join()
